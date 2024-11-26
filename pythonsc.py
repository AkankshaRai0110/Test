from datetime import datetime
from typing import Pattern, Text
import cx_Oracle
import openpyxl
from openpyxl.styles.borders import Border
from openpyxl.styles.colors import Color
from openpyxl.utils.cell import get_column_letter
import xlsxwriter
from xlsxwriter import worksheet
from configparser import ConfigParser

cx_Oracle.init_oracle_client(lib_dir=r"C:\temp\Instant Client\instantclient_19_13")
# file='Config.ini'
# config= ConfigParser()
# config.read(file)
username="cmx_ors"
pwd='cmx'
dsn = cx_Oracle.makedsn("admghp02-scan.ae.ge.com", 1521, service_name="annmdp03")
con = cx_Oracle.connect(user=username, password=pwd, dsn=dsn,
                               encoding="UTF-8")
print("Connected successfully")

workbook = xlsxwriter.Workbook('sample1.xlsx')
Parameters = workbook.add_worksheet("Parameters")
Multiple_Engine_Lvl3= workbook.add_worksheet("Multiple Engine Lvl3")
MDM_Installed_not_in_DO= workbook.add_worksheet("MDM Installed not in DO")
Install= workbook.add_worksheet("Install")
IFS_DO_ARCFT_Level3= workbook.add_worksheet("IFS DO ARCFT LVL3")
Thrust_Rating= workbook.add_worksheet("Thrust Rating")
Parameters_Including_Spare_ESNs= workbook.add_worksheet("Parameters-Including Spare ESNs")
old_ESNs_for_IFSDOENGCFG= workbook.add_worksheet("old ESNs for IFSDOENGCFG")
Parameters_IFS_MDM_Comp= workbook.add_worksheet("Parameters-IFS MDM Comp")
oldESNsIFS_MDMCFG= workbook.add_worksheet("oldESNsIFS MDMCFG")
Paperclip_IFS_Comp= workbook.add_worksheet("Paperclip IFS Comp")
DO_Flag_Mismatch= workbook.add_worksheet("DO Flag Mismatch")
Paperclip_IFS_MDM_Comp= workbook.add_worksheet("Paperclip- IFS MDM Comp")
Unmerged_Tails= workbook.add_worksheet("Unmerged Tails")
Mon_rel_Diff_ESN_Tail= workbook.add_worksheet("Mon rel Diff ESN Tail")
ESN_MDM_Diag_Mon_rel= workbook.add_worksheet("ESN MDM Diag Mon rel")
Diagnostics_Downstream= workbook.add_worksheet("Diagnostics Downstream")
Aircraft_MDM_Diag_Mon_rel= workbook.add_worksheet("Aircraft MDM Diag Mon rel")
cursor = con.cursor()
print('Fetching data from database.....')
cursor.execute("select * from DQ_SP_PARAMETERS ORDER BY aged_date")

res=cursor.fetchall()
format1 = workbook.add_format({'bold': True, 'bg_color':"#95B3D7", 'border': 1})
format2 = workbook.add_format({'bg_color': '#BFBFBF'})
format3 = workbook.add_format({'border':1, 'border_color':'#000000'})
format4 = workbook.add_format({'num_format':'dd/mm/yyyy'})

date = datetime.strptime('2011-01-01', "%Y-%m-%d")
Parameters.conditional_format('A2:A100',  {'type': 'date',
                                         'criteria': 'greater than',
                                         'value': date,
                                         'format': format4})
#sheet1= Parameter 
Parameters.conditional_format("A2:R10", {"type": "formula",
                                      "criteria": '=($B2=0)',
                                       "format": format2
                                       }
)                                        
row = 1
for r in res:
         for c in range(len(r)):
            Parameters.write(0,0,"ISSUE_START_DATE",format1)
            Parameters.write(0,1,"AGED_DATE", format1)
            Parameters.write(0,2,"ESN",format1)
            Parameters.write(0,3,"OPR",format1)
            Parameters.write(0,4,"MDM_ENG_LVL3",format1)
            Parameters.write(0,5,"TAIL_NUM",format1)
            Parameters.write(0,6,"IFS_N1_MODIFIER",format1)
            Parameters.write(0,7,"DO_N1_MODIFIER",format1)
            Parameters.write(0,8,"IFS_TCC_TIMER",format1)
            Parameters.write(0,9,"DO_TCC_TIMER",format1)
            Parameters.write(0,10,"IFS_HARDWARE_CONFIG",format1)
            Parameters.write(0,11,"DO_HARDWARE_CONFIG",format1)
            Parameters.write(0,12,"IFS_CONTROL",format1)
            Parameters.write(0,13,"DO_CONTROL",format1)
            Parameters.write(0,14,"IFS_APPLN_SELECTOR",format1)
            Parameters.write(0,15,"DO_APPLN_SELECTOR",format1)
            Parameters.write(0,16,"IFS_EGT_SHUNT",format1)
            Parameters.write(0,17,"DO_EGT_SHUNT",format1)
            Parameters.set_column(0,c,10)
        
            Parameters.write(row, c, r[c],format3)
            Parameters.set_column(c,c,15)
         row = row + 1

#Sheet2=Multiple engine Lvl3 
Multiple_Engine_Lvl3.conditional_format("A2:M20", {"type": "formula",
                                      "criteria": '=($B2=0)',
                                       "format": format2
                                       }
)
Multiple_Engine_Lvl3.conditional_format('A2:A100',  {'type': 'date',
                                         'criteria': 'greater than',
                                         'value': date,
                                         'format': format4})
cursor.execute("select * from DQ_SP_MUL_ENG_LVL3 WHERE MDM_DO_BVT NOT IN ('GE-PASSPORT 20-19BB1A','PASSPORT20-19BB1A') ORDER BY AGED_DATE")
res=cursor.fetchall()
row = 1
for r in res:
         for c in range(len(r)):
                                Multiple_Engine_Lvl3.write("A1","ISSUE_START_DATE",format1)
                                Multiple_Engine_Lvl3.write(0,1,"AGED_DATE",format1)
                                Multiple_Engine_Lvl3.write(0,2,"MDM_ESN",format1)
                                Multiple_Engine_Lvl3.write(0,3,"MDM_IFS_BVT",format1)
                                Multiple_Engine_Lvl3.write(0,4,"MDM_DO_BVT",format1)
                                Multiple_Engine_Lvl3.write(0,5,"IFS_ENGINE_LVL3",format1)
                                Multiple_Engine_Lvl3.write(0,6,"DO_ENGINE_LVL3",format1)
                                Multiple_Engine_Lvl3.write(0,7,"IFS_ENG_OPR",format1)
                                Multiple_Engine_Lvl3.write(0,8,"DO_OPR",format1)
                                Multiple_Engine_Lvl3.write(0,9,"IFS_ARCFT_PROD_NUM",format1)
                                Multiple_Engine_Lvl3.write(0,10,"IFS_ARCFT_TYPE",format1)
                                Multiple_Engine_Lvl3.write(0,11,"IFS_TAIL_NUM",format1)
                                Multiple_Engine_Lvl3.write(0,12,"DO_TAIL_NUM",format1)
                                Multiple_Engine_Lvl3.write(row, c, r[c],format3)
                                Multiple_Engine_Lvl3.set_column(c,c,15)
         row = row + 1 
# Sheet3=MDM Installed Not In DO
cursor.execute("select * from DQ_SP_IFS_INSTAL_DO_SPARE_V ORDER BY aged_date")
res=cursor.fetchall()
MDM_Installed_not_in_DO.conditional_format("A2:M30", {"type": "formula",
                                      "criteria": '=($B2=0)',
                                       "format": format2
                                       }
)
MDM_Installed_not_in_DO.conditional_format('A2:A100',  {'type': 'date',
                                         'criteria': 'greater than',
                                         'value': date,
                                         'format': format4})
row = 1
for r in res:
         for c in range(len(r)):
                                MDM_Installed_not_in_DO.write("A1","ISSUE_START_DATE",format1)
                                MDM_Installed_not_in_DO.write(0,1,"AGED_DATE",format1)
                                MDM_Installed_not_in_DO.write(0,2,"MDM_OPR",format1)
                                MDM_Installed_not_in_DO.write(0,3,"MDM_PROD_NUM",format1)
                                MDM_Installed_not_in_DO.write(0,4,"MDM_TAIL",format1)
                                MDM_Installed_not_in_DO.write(0,5," MDM Registered Tail",format1)
                                MDM_Installed_not_in_DO.write(0,6,"MDM_ENG_POSITION",format1)
                                MDM_Installed_not_in_DO.write(0,7,"MDM_ESN",format1)
                                MDM_Installed_not_in_DO.write(0,8,"MDM_IFS_TAIL",format1)
                                MDM_Installed_not_in_DO.write(0,9,"MDM_IFS_PROD_NUM",format1)
                                MDM_Installed_not_in_DO.write(0,10,"MDM_DO_TAIL",format1)
                                MDM_Installed_not_in_DO.write(0,11,"MDM DO Registered Tail",format1)
                                MDM_Installed_not_in_DO.write(0,12,"MDM_ENG_LVL3",format1)
                                MDM_Installed_not_in_DO.write(row, c, r[c],format3)
                                MDM_Installed_not_in_DO.set_column(c,c,15)
         row = row + 1        
# Sheet4=Install
cursor.execute("select * from DQ_SP_INSTALL_ENG_V ORDER BY aged_date")
res=cursor.fetchall()
Install.conditional_format("A2:Y50", {"type": "formula",
                                      "criteria": '=($B2=0)',
                                       "format": format2
                                       }
)

Install.conditional_format('A2:A100',  {'type': 'date',
                                         'criteria': 'greater than',
                                         'value': date,
                                         'format': format4})
row = 1
for r in res:
         for c in range(len(r)):
                                Install.write("A1","ISSUE_START_DATE",format1)
                                Install.write(0,1,"AGED_DATE",format1)
                                Install.write(0,2,"MDM_PROD_NUM",format1)
                                Install.write(0,3,"MDM_TAIL",format1)
                                Install.write(0,4,"MDM Registered Tail",format1)
                                Install.write(0,5,"MDM_ENG_POSITION",format1)
                                Install.write(0,6,"MDM_ESN",format1)
                                Install.write(0,7,"MDM_IFS_TAIL",format1)
                                Install.write(0,8,"MDM_IFS_PROD_NUM",format1)
                                Install.write(0,9,"MDM_DO_TAIL",format1)
                                Install.write(0,10,"MDM DO Registered Tail",format1)
                                Install.write(0,11,"IFS_PROD_NUM",format1)
                                Install.write(0,12,"IFS_TAIL",format1)
                                Install.write(0,13,"IFS_AIRCRAFT_TYPE",format1)
                                Install.write(0,14,"IFS_ENGINE_POSITION",format1)
                                Install.write(0,15,"IFS_ESN",format1)
                                Install.write(0,16,"IFS_OPR",format1)
                                Install.write(0,17,"IFS_OPRTNL_STATUS",format1)
                                Install.write(0,18,"IFS_SERV_STATUS",format1)
                                Install.write(0,19,"DO_ESN",format1)
                                Install.write(0,20,"DO_TAIL",format1)
                                Install.write(0,21,"DO_ENGINE_POSITION",format1)
                                Install.write(0,22,"Diagnostics Monitoring cd",format1)
                                Install.write(0,23,"DO Registered Tail",format1)
                                Install.write(0,24,"DO_STATUS",format1)
                                Install.write(row, c, r[c],format3)
                                Install.set_column(c,c,15)
         row = row + 1 
#Sheet5=IFS DO ACRFT LVL3
cursor.execute("select * from DQ_SP_IFS_DO_AIRC_TYPE_V ORDER BY aged_date")
res=cursor.fetchall()
IFS_DO_ARCFT_Level3.conditional_format('A2:A100',  {'type': 'date',
                                         'criteria': 'greater than',
                                         'value': date,
                                         'format': format4})
row = 1
for r in res:
         for c in range(len(r)):
                                IFS_DO_ARCFT_Level3.write("A1","ISSUE_START_DATE",format1)
                                IFS_DO_ARCFT_Level3.write(0,1,"AGED_DATE",format1)
                                IFS_DO_ARCFT_Level3.write(0,2,"MDM_IFS_TAIL",format1)
                                IFS_DO_ARCFT_Level3.write(0,3,"ACRFT_PRODTN_NUM",format1)
                                IFS_DO_ARCFT_Level3.write(0,4,"TAIL_NUM",format1)
                                IFS_DO_ARCFT_Level3.write(0,5,"PART_NUM",format1)
                                IFS_DO_ARCFT_Level3.write(0,6,"DO_TAIL",format1)
                                IFS_DO_ARCFT_Level3.write(0,7,"MDO Registered TailL",format1)
                                IFS_DO_ARCFT_Level3.write(0,8,"DO_LVL3",format1)
                                IFS_DO_ARCFT_Level3.write(row, c, r[c],format3)
                                IFS_DO_ARCFT_Level3.set_column(c,c,15)
         row = row + 1

#("Paperclip IFS Comp")
cursor.execute("select * from  cmx_ors.DQ_SC_ORG_P2_P3")
res=cursor.fetchall()
row = 1
for r in res:
         for c in range(len(r)):
                                Paperclip_IFS_Comp.write("A1","ICAO_CD",format1)
                                Paperclip_IFS_Comp.write(0,1,"P2_ORG_NAME",format1)
                                Paperclip_IFS_Comp.write(0,2,"P3_ORG_NAME",format1)
                                Paperclip_IFS_Comp.write(0,3,"P2_ORG_TYPE",format1)
                                Paperclip_IFS_Comp.write(0,4,"P3_ORG_TYPE",format1)
                                Paperclip_IFS_Comp.write(0,5,"P2_LNG",format1)
                                Paperclip_IFS_Comp.write(0,6,"P3_LNG",format1)
                                Paperclip_IFS_Comp.write(0,7,"P2_ORG_CATEGORY",format1)
                                Paperclip_IFS_Comp.write(0,8,"P3_ORG_CATEGORY",format1)
                                Paperclip_IFS_Comp.write(0,9,"P2_FSE_CVG",format1)
                                Paperclip_IFS_Comp.write(0,10,"P3_FSE_CVG",format1)
                                Paperclip_IFS_Comp.write(row, c, r[c],format3)
                                Paperclip_IFS_Comp.set_column(c,c,15)
         row = row + 1

#("DO Flag Mismatch")
cursor.execute("select * from DQ_SC_DO_FLAG_MISMATCH")
res=cursor.fetchall()
row = 1
for r in res:
         for c in range(len(r)):
                                DO_Flag_Mismatch.write("A1","MDM_TAIL",format1)
                                DO_Flag_Mismatch.write(0,1,"IFS_TAIL_NUM",format1)
                                DO_Flag_Mismatch.write(0,2,"MDM_INCOMING_TAIL",format1)
                                DO_Flag_Mismatch.write(0,3,"MDM_AIRCRAFT_PROD_NUM",format1)
                                DO_Flag_Mismatch.write(0,4,"IFS_OWNER",format1)
                                DO_Flag_Mismatch.write(0,5,"IFS_OPERATOR",format1)
                                DO_Flag_Mismatch.write(0,6,"MDM_MONITORING_CUSTOMER",format1)
                                DO_Flag_Mismatch.write(0,7,"ENGINE_POS_NUM",format1)
                                DO_Flag_Mismatch.write(0,8,"MDM_ESN",format1)
                                DO_Flag_Mismatch.write(0,9,"MDM_DIAGNOSTIC_DOWNSTREAM",format1)
                                DO_Flag_Mismatch.write(0,10,"DIAGNOSTIC_TAIL",format1)
                                DO_Flag_Mismatch.write(0,11,"DIAGNOSTICS_INCMG_TAIL",format1)
                                DO_Flag_Mismatch.write(0,12,"DIAGNOSTICS_CUSTOMER",format1)
                                DO_Flag_Mismatch.write(0,13,"DIAGNOSTIC_ESN",format1)
                                DO_Flag_Mismatch.write(0,14,"DO_XREF_EXISTS",format1)
                                DO_Flag_Mismatch.write(row, c, r[c],format3)
                                DO_Flag_Mismatch.set_column(c,c,15)
         row = row + 1

#("Paperclip- IFS MDM Comp")

cursor.execute("select * from  cmx_ors.DQ_SC_IM_PAPERCLIP_8_8")
res=cursor.fetchall()
row = 1
for r in res:
         for c in range(len(r)):
                                Paperclip_IFS_MDM_Comp.write("A1","MDM_ICAO_CD",format1)
                                Paperclip_IFS_MDM_Comp.write(0,1,"MDM_ORG_NM",format1)
                                Paperclip_IFS_MDM_Comp.write(0,2,"MDM_FSE_CVG",format1)
                                Paperclip_IFS_MDM_Comp.write(0,3,"IFS_FSE_CVG",format1)
                                Paperclip_IFS_MDM_Comp.write(0,4,"MDM_LNG",format1)
                                Paperclip_IFS_MDM_Comp.write(0,5,"IFS_LNG",format1)
                                Paperclip_IFS_MDM_Comp.write(0,6,"IFS_ORG_CATEGORY",format1)
                                Paperclip_IFS_MDM_Comp.write(0,7,"MDM_ORG_CATEGORY",format1)
                                Paperclip_IFS_MDM_Comp.write(0,8,"IFS_ORG_SUBTYPE",format1)
                                Paperclip_IFS_MDM_Comp.write(0,9,"MDM_ORG_SUBTYPE",format1)
                                Paperclip_IFS_MDM_Comp.write(0,10,"DSPLY_NM",format1)
                                Paperclip_IFS_MDM_Comp.write(row, c, r[c],format3)
                                Paperclip_IFS_MDM_Comp.set_column(c,c,15)
         row = row + 1

cursor.execute("select * from dq_airc_unmerged_moni_rel")
res=cursor.fetchall()
row = 1
for r in res:
         for c in range(len(r)):
                                Unmerged_Tails.write("A1","Diag. Tail",format1)
                                Unmerged_Tails.write(0,1,"Diag. Reg Tail",format1)
                                Unmerged_Tails.write(0,2," Diag. Customer",format1)
                                Unmerged_Tails.write(0,3,"COMMENTS",format1)
                                Unmerged_Tails.write(0,4,"IFS Reg. Tail",format1)
                                Unmerged_Tails.write(row, c, r[c],format3)
                                Unmerged_Tails.set_column(c,c,15)
         row = row + 1
cursor.execute("select * from dq_esn_airc_diff_moni_rel")
res=cursor.fetchall()
row = 1
for r in res:
         for c in range(len(r)):
                                Mon_rel_Diff_ESN_Tail.write("A1","MDM ESN",format1)
                                Mon_rel_Diff_ESN_Tail.write(0,1,"MDM Monitoring rel",format1)
                                Mon_rel_Diff_ESN_Tail.write(0,2,"MDM TAIL",format1)
                                Mon_rel_Diff_ESN_Tail.write(0,3,"TAIL Monitoring rel",format1)
                                Mon_rel_Diff_ESN_Tail.write(row, c, r[c],format3)
                                Mon_rel_Diff_ESN_Tail.set_column(c,c,15)
         row = row + 1
cursor.execute("select * from dq_eng_diff_moni_rel")
res=cursor.fetchall()
row = 1
for r in res:
         for c in range(len(r)):
                                ESN_MDM_Diag_Mon_rel.write("A1","ESN",format1)
                                ESN_MDM_Diag_Mon_rel.write(0,1,"MDM Monitoring rel",format1)
                                ESN_MDM_Diag_Mon_rel.write(0,2,"Diag. Customer",format1)
                                ESN_MDM_Diag_Mon_rel.write(row, c, r[c],format3)
                                ESN_MDM_Diag_Mon_rel.set_column(c,c,15)
         row = row + 1 
cursor.execute("select * from dq_eng_not_enabled_moni_rel")
res=cursor.fetchall()
row = 1
for r in res:
         for c in range(len(r)):
                                Diagnostics_Downstream.write("A1","ESN",format1)
                                Diagnostics_Downstream.write(0,1,"Diagnostics Customer",format1)
                                Diagnostics_Downstream.write(0,2,"Diagnostics Downstream",format1)
                                Diagnostics_Downstream.write(row, c, r[c],format3)
                                Diagnostics_Downstream.set_column(c,c,15)
         row = row + 1 
cursor.execute("select * from dq_airc_diff_moni_rel")
res=cursor.fetchall()
row = 1
for r in res:
         for c in range(len(r)):
                                Aircraft_MDM_Diag_Mon_rel.write("A1","Diag Tail Num",format1)
                                Aircraft_MDM_Diag_Mon_rel.write(0,1,"MDM Monitoring rel",format1)
                                Aircraft_MDM_Diag_Mon_rel.write(0,2,"Diag. Customer",format1)
                                Aircraft_MDM_Diag_Mon_rel.write(row, c, r[c],format3)
                                Aircraft_MDM_Diag_Mon_rel.set_column(c,c,15)
                                
         row = row + 1 

cursor.execute("select * from DQ_SC_ENGINE_THRUST_RTG_V")
res=cursor.fetchall()
row = 1
for r in res:
         for c in range(len(r)):
                                Thrust_Rating.write("A1","MDM_ESN",format1)
                                Thrust_Rating.write(0,1,"MDM_ENGINE_LVL2",format1)
                                Thrust_Rating.write(0,2,"IFS_THRUST",format1)
                                Thrust_Rating.write(0,3,"MDM_THRUST",format1)
                                Thrust_Rating.write(0,4,"DO_THRUST",format1)
                                Thrust_Rating.write(0,5,"IFS_OPRNTL_STATUS",format1)
                                Thrust_Rating.write(0,6,"IFS_SERVC_STAUS",format1)
                                Thrust_Rating.write(0,7,"DO_STATUS",format1)
                                Thrust_Rating.write(0,8,"Diagnostics Monitoring cd",format1)
                                Thrust_Rating.write(0,9,"IFS_OPR",format1)
                                Thrust_Rating.write(0,10,"IFS_OWNER",format1)
                                Thrust_Rating.write(row, c, r[c],format3)
                                Thrust_Rating.set_column(c,c,15)
                                
         row = row + 1 

#("Parameters-Including Spare ESNs")
cursor.execute("SELECT a.*,b.status mdm_status ,c.status IFS_status, d.status DO_status FROM DQ_SC_IFS_DO_ENG_CFG_10_24_V a LEFT JOIN c_bo_serialized_engine b ON a.esn=b.engine_serial_num LEFT JOIN c_bo_serialized_engine_xref c ON a.esn          =c.engine_serial_num AND c.rowid_system='IFS' LEFT JOIN c_bo_serialized_engine_xref d ON a.esn =d.engine_serial_num AND d.rowid_system='DO'")
res=cursor.fetchall()
row = 1
for r in res:
         for c in range(len(r)):
                                Parameters_Including_Spare_ESNs.write("A1","ESN",format1)
                                Parameters_Including_Spare_ESNs.write(0,1,"IFS_N1",format1)
                                Parameters_Including_Spare_ESNs.write(0,2,"DO_N1",format1)
                                Parameters_Including_Spare_ESNs.write(0,3,"IFS_APP",format1)
                                Parameters_Including_Spare_ESNs.write(0,4,"DO_APP",format1)
                                Parameters_Including_Spare_ESNs.write(0,5,"IFS_CFG",format1)
                                Parameters_Including_Spare_ESNs.write(0,6,"DO_CFG",format1)
                                Parameters_Including_Spare_ESNs.write(0,7,"IFS_CNT",format1)
                                Parameters_Including_Spare_ESNs.write(0,8,"DO_CNT",format1)
                                Parameters_Including_Spare_ESNs.write(0,9,"IFS_TCC",format1)
                                Parameters_Including_Spare_ESNs.write(0,10,"DO_TCC",format1)
                                Parameters_Including_Spare_ESNs.write(0,11,"IFS_OPR",format1)
                                Parameters_Including_Spare_ESNs.write(0,12,"Diagnostics Monitoring cd",format1)
                                Parameters_Including_Spare_ESNs.write(0,13,"MDM_STATUS",format1)
                                Parameters_Including_Spare_ESNs.write(0,14,"IFS_STATUS",format1)
                                Parameters_Including_Spare_ESNs.write(0,15,"DO_STATUS",format1)
                                Parameters_Including_Spare_ESNs.write(0,17,"",format1)
                                
                                Parameters_Including_Spare_ESNs.write(row, c, r[c],format3)
                                Parameters_Including_Spare_ESNs.set_column(c,c,15)
                                
         row = row + 1 
#sheet7.autofilter_ref='A1:S1'

old_ESNs_for_IFSDOENGCFG.write('A1','ESN', format1)
cursor.execute("select * from  cmx_ors.DQ_SC_IM_ENGINE_CONFIG")
res=cursor.fetchall()
row = 1
for r in res:
         for c in range(len(r)):
                                Parameters_IFS_MDM_Comp.write("A1","MDM_ESN",format1)
                                Parameters_IFS_MDM_Comp.write(0,1,"MDM_APP_SEL",format1)
                                Parameters_IFS_MDM_Comp.write(0,2,"MDM_APP_SEL_DT",format1)
                                Parameters_IFS_MDM_Comp.write(0,3,"IFS_APP_SEL",format1)
                                Parameters_IFS_MDM_Comp.write(0,4,"IFS_APP_SEL_DT",format1)
                                Parameters_IFS_MDM_Comp.write(0,5,"MDM_CNTRL",format1)
                                Parameters_IFS_MDM_Comp.write(0,6,"MDM_CNTRL_DT",format1)
                                Parameters_IFS_MDM_Comp.write(0,7,"IFS_CNTRL",format1)
                                Parameters_IFS_MDM_Comp.write(0,8,"IFS_CNTRL_DT",format1)
                                Parameters_IFS_MDM_Comp.write(0,9,"MDM_N1",format1)
                                Parameters_IFS_MDM_Comp.write(0,10,"MDM_N1_DT",format1)
                                Parameters_IFS_MDM_Comp.write(0,11,"IFS_N1",format1)
                                Parameters_IFS_MDM_Comp.write(0,12,"IFS_N1_DT",format1)
                                Parameters_IFS_MDM_Comp.write(0,13,"MDM_TCC",format1)
                                Parameters_IFS_MDM_Comp.write(0,14,"MDM_TCC_DT",format1)
                                Parameters_IFS_MDM_Comp.write(0,15,"IFS_TCC",format1)
                                Parameters_IFS_MDM_Comp.write(0,16,"IFS_TCC_DT",format1)
                                Parameters_IFS_MDM_Comp.write(0,17,"MDM_HRDW_CFG",format1)
                                Parameters_IFS_MDM_Comp.write(0,18,"MDM_HRDW_CFG_DT",format1)
                                Parameters_IFS_MDM_Comp.write(0,19,"IFS_HRDW_CFG",format1)
                                Parameters_IFS_MDM_Comp.write(0,20,"IFS_HRDW_CFG_DT",format1)
                                Parameters_IFS_MDM_Comp.write(0,23,"",format1)
                                Parameters_IFS_MDM_Comp.write(row, c, r[c],format3)
                                Parameters_IFS_MDM_Comp.set_column(c,c,15)

         row = row + 1 
#sheet9.autofilter_ref='A1:X1'

oldESNsIFS_MDMCFG.write('A1','ESN', format1)
workbook.close()
cursor.close()

print("Data has been fetched and stored in Excel.")
print("Applying specific formats for Sheets")
print("Please wait till formtting completed....")

from openpyxl import load_workbook
src_wb = load_workbook('sample1.xlsx')
dest_wb = load_workbook('destination.xlsx')

dest_wb.create_sheet("Thrust")
dest_wb.create_sheet("Parameter-1")
dest_wb.create_sheet("Parameter-2")

Thrust_Rating_sheet = src_wb['Thrust Rating']
Thrust_sheet = dest_wb['Thrust']

for i in range(1, Thrust_Rating_sheet.max_row+1):
    for j in range(1, Thrust_Rating_sheet.max_column+1):
        Thrust_sheet.cell(row=i, column=j).value = Thrust_Rating_sheet.cell(row=i, column=j).value

src_parameter1_sheet = src_wb['Parameters-Including Spare ESNs']
dest_parameter1_sheet = dest_wb['Parameter-1']

for i in range(1, src_parameter1_sheet.max_row+1):
    for j in range(1, src_parameter1_sheet.max_column+1):
        dest_parameter1_sheet.cell(row=i, column=j).value = src_parameter1_sheet.cell(row=i, column=j).value

src_parameter2_sheet = src_wb['Parameters-IFS MDM Comp']
dest_parameter2_sheet = dest_wb['Parameter-2']

for i in range(1, src_parameter2_sheet.max_row+1):
    for j in range(1, src_parameter2_sheet.max_column+1):
        dest_parameter2_sheet.cell(row=i, column=j).value = src_parameter2_sheet.cell(row=i, column=j).value

src_wb.save('sample1.xlsx')
dest_wb.save('destination.xlsx')

from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
wb = load_workbook('Production_3-way_Scorecard_20230502.xlsx')
wb2 = load_workbook('sample1.xlsx')

thin = Side(border_style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

sheet1 = wb['Parameters-Including Spare ESNs']
sheet8 = wb2['old ESNs for IFSDOENGCFG']

for i in range(1,sheet1.max_row+1):
    for j in range(1,sheet1.max_column):
        sheet8.cell(row=i,column=1).value = sheet1.cell(row=i,column=1).value
        sheet8.cell(row=i, column=1).border = border

wb2.save('sample1.xlsx')
sheet1 = wb['Parameters-IFS MDM Comp']
sheet10 = wb2['oldESNsIFS MDMCFG']

for i in range(1,sheet1.max_row+1):
    for j in range(1,sheet1.max_column):
        sheet10.cell(row=i,column=1).value = sheet1.cell(row=i,column=1).value
        sheet10.cell(row=i, column=1).border = border

wb2.save('sample1.xlsx')


from openpyxl import load_workbook, Workbook
import xlwings as xw
import pandas as pd
from pathlib import Path

DATA_DIR = Path.cwd() / 'python'

df = pd.DataFrame()
path =r'C:\Users\2009954\TCS GE Confidential\pythoncode\sample1.xlsx'
writer = pd.ExcelWriter(path, engine='openpyxl',mode='r+', if_sheet_exists='replace')
writer.book = load_workbook(path)
writer.sheets = dict((ws.title,ws) for ws in writer.book.worksheets)
orders = pd.read_excel("sample1.xlsx",sheet_name = 'Parameters-Including Spare ESNs',engine='openpyxl')
returns = pd.read_excel("sample1.xlsx",sheet_name = 'old ESNs for IFSDOENGCFG',engine='openpyxl')

df['exists']= (orders.ESN).isin(returns.ESN)
df.to_excel(writer, sheet_name='Parameters-Including Spare ESNs',startcol=18, index = False)
Parameters_Including_Spare_ESNs.conditional_format("A2:S100", {"type": "formula",
                                      "criteria": '=($S2="FALSE")',
                                       "format": format2
                                       }
)

writer.save()

from openpyxl import load_workbook, Workbook
import xlwings as xw
import pandas as pd
from pathlib import Path

DATA_DIR = Path.cwd() / 'python'

df = pd.DataFrame()
path =r'C:\Users\2009954\TCS GE Confidential\pythoncode\sample1.xlsx'
writer = pd.ExcelWriter(path, engine='openpyxl', mode='r+', if_sheet_exists= 'replace')
writer.book = load_workbook(path)
writer.sheets = dict((ws.title,ws) for ws in writer.book.worksheets)
orders = pd.read_excel("sample1.xlsx",sheet_name = 'Parameters-IFS MDM Comp',engine='openpyxl')
returns = pd.read_excel("sample1.xlsx",sheet_name = 'oldESNsIFS MDMCFG',engine='openpyxl')

df['exists']= (orders.MDM_ESN).isin(returns.MDM_ESN)
df.to_excel(writer, sheet_name='Parameters-IFS MDM Comp',startcol=24, index = False)
writer.save()

from openpyxl import load_workbook, Workbook
import xlwings as xw
import pandas as pd
from pathlib import Path

DATA_DIR = Path.cwd() / 'python'

df = pd.DataFrame()
path =r'C:\Users\2009954\TCS GE Confidential\pythoncode\sample1.xlsx'
writer = pd.ExcelWriter(path, engine='openpyxl', mode='r+', if_sheet_exists='replace')
writer.book = load_workbook(path)
writer.sheets = dict((ws.title,ws) for ws in writer.book.worksheets)
orders = pd.read_excel("sample1.xlsx",sheet_name = 'Thrust Rating',engine='openpyxl')
returns = pd.read_excel("Production_3-way_Scorecard_20230502.xlsx",sheet_name = 'Thrust Rating',engine='openpyxl')

df['exists']= (orders.MDM_ESN).isin(returns.MDM_ESN)
df.to_excel(writer, sheet_name='Thrust Rating',startcol=12, index = False)
Thrust_Rating.conditional_format("A2:M100", {"type": "formula",
                                      "criteria": '=($M2="FALSE")',
                                       "format": format2
                                       }
)

writer.save()

import openpyxl
from openpyxl import load_workbook 
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.styles import Font, Border, Side, PatternFill
src_wb = load_workbook('destination.xlsx')
dest_wb = load_workbook('sample1.xlsx')

thin = Side(border_style="thin", color="000000")
bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

src_thrust_sheet = src_wb['Thrust']
dest_thrust_sheet = dest_wb['Thrust Rating']

dim_holder = DimensionHolder(worksheet=dest_thrust_sheet)

for i in range(1, src_thrust_sheet.max_row+1):
    for j in range(1, src_thrust_sheet.max_column):
        dest_thrust_sheet.cell(row=i, column=j).value = src_thrust_sheet.cell(row=i, column=j).value
        cell_header = dest_thrust_sheet.cell(1, j)
        cell_bdr = dest_thrust_sheet.cell(i,j)
        cell_header.fill=PatternFill(start_color='95B3D7', end_color='95B3D7', fill_type='solid')
        cell_header.font= Font(bold=True)
        cell_bdr.border= bdr
        dim_holder[get_column_letter(j)] =ColumnDimension(dest_thrust_sheet, min=j, max=j, width=10)

dest_thrust_sheet.column_dimensions=dim_holder
#ColumnDimension(dest_thrust_sheet, bestFit=True, width=20)
        
        

src_parameter1_sheet = src_wb['Parameter-1']
dest_parameter1_sheet = dest_wb['Parameters-Including Spare ESNs']
dest_parameter1_sheet.auto_filter.ref='A1:S1'

for i in range(1, src_parameter1_sheet.max_row+1):
    for j in range(1, src_parameter1_sheet.max_column+1):
        dest_parameter1_sheet.cell(row=i, column=j).value = src_parameter1_sheet.cell(row=i, column=j).value
        cell_header = dest_parameter1_sheet.cell(1, j)
        cell_bdr = dest_parameter1_sheet.cell(i,j)
        cell_header.fill=PatternFill(start_color='95B3D7', end_color='95B3D7', fill_type='solid')
        cell_header.font= Font(bold=True)
        cell_bdr.border= bdr

src_parameter2_sheet = src_wb['Parameter-2']
dest_parameter2_sheet = dest_wb['Parameters-IFS MDM Comp']
dest_parameter2_sheet.auto_filter.ref='A1:X1'

for i in range(1, src_parameter2_sheet.max_row+1):
    for j in range(1, src_parameter2_sheet.max_column+1):
        dest_parameter2_sheet.cell(row=i, column=j).value = src_parameter2_sheet.cell(row=i, column=j).value
        cell_header = dest_parameter2_sheet.cell(1, j)
        cell_bdr = dest_parameter2_sheet.cell(i,j)
        cell_header.fill=PatternFill(start_color='95B3D7', end_color='95B3D7', fill_type='solid')
        cell_header.font= Font(bold=True)
        cell_bdr.border= bdr
        cell_bdr

src_wb.save('destination.xlsx')
dest_wb.save('sample1.xlsx')

src_wb.remove(src_parameter2_sheet)
src_wb.remove(src_parameter1_sheet)
src_wb.remove(src_thrust_sheet)

print("Scorecard has been created !")

